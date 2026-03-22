"""
文件处理器组件
基于marker组件实现的文件到markdown转换器
支持多种文件格式，提供高质量的markdown输出
"""

import os
import asyncio
import uuid
from typing import Dict, Any, Optional, Union, List
from pathlib import Path
from fastapi import UploadFile, HTTPException

from .config import MarkerConfig, ConfigManager
from .markdown_extractor import MarkdownExtractor, MarkdownContent


class FileProcessor:
    """文件处理器主类"""
    
    def __init__(self, config: Optional[MarkerConfig] = None, upload_dir: str = "uploads"):
        self.config_manager = ConfigManager(config)
        self.markdown_extractor = MarkdownExtractor()
        self.upload_dir = Path(upload_dir)
        self.upload_dir.mkdir(exist_ok=True)
        
        # 初始化marker组件（延迟导入以避免依赖问题）
        self._converter = None
        self._model_dict = None
        
    def _init_marker_components(self):
        """初始化marker组件"""
        if self._converter is None:
            try:
                from marker.converters.pdf import PdfConverter
                from marker.models import create_model_dict
                from marker.config.parser import ConfigParser

                # 创建配置解析器
                config_dict = self.config_manager.get_config_dict()
                config_parser = ConfigParser(config_dict)

                # 创建模型字典
                self._model_dict = create_model_dict()

                # 创建转换器
                self._converter = PdfConverter(
                    config=config_dict,
                    artifact_dict=self._model_dict,
                    processor_list=config_parser.get_processors(),
                    renderer=config_parser.get_renderer(),
                )

                # 如果使用LLM，设置LLM服务
                if config_dict.get('use_llm', False):
                    llm_service = config_parser.get_llm_service()
                    if llm_service:
                        self._converter.llm_service = llm_service

                self._marker_available = True

            except ImportError as e:
                print(f"⚠️ Marker依赖不可用，将使用基础文本提取: {str(e)}")
                self._marker_available = False
                self._converter = None
            except Exception as e:
                print(f"⚠️ Marker初始化失败，将使用基础文本提取: {str(e)}")
                self._marker_available = False
                self._converter = None
    
    async def process_file(self, file_path: Union[str, Path]) -> MarkdownContent:
        """处理单个文件"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if not self.config_manager.is_file_supported(file_path.name):
            raise ValueError(f"不支持的文件类型: {file_path.suffix}")
        
        # 验证文件大小
        file_size = file_path.stat().st_size
        if file_size > self.config_manager.config.max_file_size:
            max_size_mb = self.config_manager.config.max_file_size / 1024 / 1024
            raise ValueError(f"文件大小超过限制 ({max_size_mb:.1f}MB)")
        
        try:
            # 初始化marker组件
            self._init_marker_components()

            # 检查marker是否可用
            if hasattr(self, '_marker_available') and self._marker_available and self._converter:
                # 使用marker转换文件
                rendered_result = await self._convert_with_marker(str(file_path))
                # 提取markdown内容
                markdown_content = self.markdown_extractor.extract_from_rendered(rendered_result)
            else:
                # 使用基础文本提取作为后备方案
                markdown_content = await self._fallback_text_extraction(file_path)

            return markdown_content

        except Exception as e:
            # 如果marker失败，尝试基础提取
            try:
                print(f"⚠️ Marker处理失败，尝试基础文本提取: {str(e)}")
                markdown_content = await self._fallback_text_extraction(file_path)
                return markdown_content
            except Exception as fallback_error:
                raise Exception(f"文件处理失败 - Marker: {str(e)}, 基础提取: {str(fallback_error)}")
    
    async def _convert_with_marker(self, file_path: str):
        """使用marker转换文件"""
        try:
            # 在线程池中运行marker转换（因为marker是同步的）
            loop = asyncio.get_event_loop()
            rendered_result = await loop.run_in_executor(
                None,
                self._converter,
                file_path
            )
            return rendered_result
        except Exception as e:
            raise Exception(f"Marker转换失败: {str(e)}")

    async def _fallback_text_extraction(self, file_path: Path):
        """基础文本提取作为后备方案"""
        from .markdown_extractor import MarkdownContent

        file_extension = file_path.suffix.lower()
        content = ""

        try:
            if file_extension == '.pdf':
                content = await self._extract_pdf_basic(file_path)
            elif file_extension in ['.docx', '.doc']:
                content = await self._extract_docx_basic(file_path)
            elif file_extension == '.txt':
                content = await self._extract_txt_basic(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                content = await self._extract_xlsx_basic(file_path)
            else:
                content = f"文件类型 {file_extension} 需要marker-pdf支持，请安装: pip install marker-pdf"

            # 创建基础的MarkdownContent对象
            return MarkdownContent(
                text=content,
                images={},
                metadata={"extraction_method": "basic", "file_type": file_extension},
                tables=[],
                headers=[],
                links=[],
                code_blocks=[],
                math_expressions=[]
            )

        except Exception as e:
            raise Exception(f"基础文本提取失败: {str(e)}")

    async def _extract_pdf_basic(self, file_path: Path) -> str:
        """基础PDF文本提取"""
        try:
            import PyPDF2
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                content = []
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    content.append(page.extract_text())
                return '\n'.join(content)
        except ImportError:
            return "PDF处理需要PyPDF2或marker-pdf，请安装: pip install PyPDF2 或 pip install marker-pdf"
        except Exception as e:
            raise Exception(f"PDF提取失败: {str(e)}")

    async def _extract_docx_basic(self, file_path: Path) -> str:
        """基础DOCX文本提取"""
        try:
            from docx import Document
            doc = Document(file_path)
            content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    content.append(paragraph.text)
            return '\n'.join(content)
        except ImportError:
            return "DOCX处理需要python-docx，请安装: pip install python-docx"
        except Exception as e:
            raise Exception(f"DOCX提取失败: {str(e)}")

    async def _extract_txt_basic(self, file_path: Path) -> str:
        """基础TXT文本提取"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='gbk') as f:
                    return f.read()
            except Exception as e:
                raise Exception(f"文本文件编码错误: {str(e)}")

    async def _extract_xlsx_basic(self, file_path: Path) -> str:
        """基础XLSX文本提取"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"工作表: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(row_data):
                        content.append('\t'.join(row_data))
                content.append('')
            return '\n'.join(content)
        except ImportError:
            return "Excel处理需要openpyxl，请安装: pip install openpyxl"
        except Exception as e:
            raise Exception(f"Excel提取失败: {str(e)}")
    
    async def process_upload_file(self, file: UploadFile) -> Dict[str, Any]:
        """处理上传的文件"""
        # 验证文件类型
        if not self.config_manager.is_file_supported(file.filename):
            supported_types = ', '.join(self.config_manager.get_supported_extensions())
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件类型: {Path(file.filename).suffix}。支持的类型: {supported_types}"
            )
        
        # 验证文件大小
        file_content = await file.read()
        if len(file_content) > self.config_manager.config.max_file_size:
            max_size_mb = self.config_manager.config.max_file_size / 1024 / 1024
            raise HTTPException(
                status_code=400,
                detail=f"文件大小超过限制 ({max_size_mb:.1f}MB)"
            )
        
        # 重置文件指针
        await file.seek(0)
        
        # 生成唯一文件名
        file_extension = Path(file.filename).suffix.lower()
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = self.upload_dir / unique_filename
        
        try:
            # 保存文件
            with open(file_path, 'wb') as f:
                f.write(file_content)
            
            # 处理文件
            markdown_content = await self.process_file(file_path)
            
            # 获取结构化内容
            structured_content = self.markdown_extractor.get_structured_content(markdown_content)
            
            return {
                "status": "success",
                "filename": file.filename,
                "saved_filename": unique_filename,
                "file_path": str(file_path),
                "file_size": len(file_content),
                "file_type": file_extension,
                "markdown_content": structured_content,
                "processing_config": self.config_manager.get_config_dict()
            }
            
        except Exception as e:
            # 如果处理失败，删除已保存的文件
            if file_path.exists():
                file_path.unlink()
            raise HTTPException(status_code=500, detail=f"文件处理失败: {str(e)}")
    
    async def process_multiple_files(self, file_paths: List[Union[str, Path]]) -> List[Dict[str, Any]]:
        """批量处理多个文件"""
        results = []
        
        for file_path in file_paths:
            try:
                markdown_content = await self.process_file(file_path)
                structured_content = self.markdown_extractor.get_structured_content(markdown_content)
                
                results.append({
                    "status": "success",
                    "file_path": str(file_path),
                    "markdown_content": structured_content
                })
                
            except Exception as e:
                results.append({
                    "status": "error",
                    "file_path": str(file_path),
                    "error": str(e)
                })
        
        return results
    
    def update_config(self, **kwargs) -> None:
        """更新配置"""
        self.config_manager.update_config(**kwargs)
        # 重置converter以使用新配置
        self._converter = None
    
    def get_config(self) -> Dict[str, Any]:
        """获取当前配置"""
        return self.config_manager.get_config_dict()
    
    def search_in_content(self, markdown_content: MarkdownContent, query: str, case_sensitive: bool = False) -> List[Dict[str, Any]]:
        """在内容中搜索"""
        return self.markdown_extractor.search_content(markdown_content, query, case_sensitive)
    
    def save_content(self, markdown_content: MarkdownContent, output_path: str) -> None:
        """保存内容到文件"""
        self.markdown_extractor.save_content(markdown_content, output_path)
    
    def get_file_info(self, file_path: str) -> Optional[Dict[str, Any]]:
        """获取文件信息"""
        try:
            path = Path(file_path)
            if not path.exists():
                return None
            
            stat = path.stat()
            return {
                "filename": path.name,
                "file_size": stat.st_size,
                "file_extension": path.suffix.lower(),
                "is_supported": self.config_manager.is_file_supported(path.name),
                "created_time": stat.st_ctime,
                "modified_time": stat.st_mtime
            }
        except Exception:
            return None
    
    def delete_file(self, file_path: str) -> bool:
        """删除文件"""
        try:
            path = Path(file_path)
            if path.exists():
                path.unlink()
                return True
            return False
        except Exception:
            return False
    
    def get_supported_formats(self) -> Dict[str, List[str]]:
        """获取支持的文件格式信息"""
        extensions = self.config_manager.get_supported_extensions()
        
        format_groups = {
            "文档": ['.pdf', '.docx', '.doc', '.html', '.htm', '.epub'],
            "演示文稿": ['.pptx', '.ppt'],
            "电子表格": ['.xlsx', '.xls'],
            "图片": ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif']
        }
        
        supported_formats = {}
        for group, group_extensions in format_groups.items():
            supported_formats[group] = [ext for ext in group_extensions if ext in extensions]
        
        return supported_formats


class AsyncFileProcessor(FileProcessor):
    """异步文件处理器"""
    
    def __init__(self, config: Optional[MarkerConfig] = None, upload_dir: str = "uploads", max_concurrent: int = 3):
        super().__init__(config, upload_dir)
        self.max_concurrent = max_concurrent
        self._semaphore = asyncio.Semaphore(max_concurrent)
    
    async def process_file_with_limit(self, file_path: Union[str, Path]) -> MarkdownContent:
        """带并发限制的文件处理"""
        async with self._semaphore:
            return await self.process_file(file_path)
    
    async def process_multiple_files_concurrent(self, file_paths: List[Union[str, Path]]) -> List[Dict[str, Any]]:
        """并发处理多个文件"""
        tasks = []
        
        for file_path in file_paths:
            task = asyncio.create_task(self._process_single_file_safe(file_path))
            tasks.append(task)
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        processed_results = []
        for i, result in enumerate(results):
            if isinstance(result, Exception):
                processed_results.append({
                    "status": "error",
                    "file_path": str(file_paths[i]),
                    "error": str(result)
                })
            else:
                processed_results.append(result)
        
        return processed_results
    
    async def _process_single_file_safe(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """安全地处理单个文件"""
        try:
            markdown_content = await self.process_file_with_limit(file_path)
            structured_content = self.markdown_extractor.get_structured_content(markdown_content)
            
            return {
                "status": "success",
                "file_path": str(file_path),
                "markdown_content": structured_content
            }
        except Exception as e:
            return {
                "status": "error",
                "file_path": str(file_path),
                "error": str(e)
            }
